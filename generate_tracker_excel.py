"""
AI Software Engineer OS - 10 Week Roadmap Tracker
Generates a polished, production-grade .xlsx file with 10 fully-populated sheets.
"""
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# =========================================================================
# DESIGN SYSTEM
# =========================================================================
HEADER_FILL = PatternFill("solid", fgColor="3730A3")  # deep indigo
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=18, bold=True, color="1E1B4B")
SUBTITLE_FONT = Font(name="Calibri", size=10, italic=True, color="6B7280")
BAND_FILL = PatternFill("solid", fgColor="F9FAFB")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
BORDER_COLOR = "E5E7EB"

thin = Side(border_style="thin", color=BORDER_COLOR)
CELL_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

# KPI card palette
ACCENT_INDIGO = PatternFill("solid", fgColor="EEF2FF")
ACCENT_GREEN  = PatternFill("solid", fgColor="D1FAE5")
ACCENT_AMBER  = PatternFill("solid", fgColor="FEF3C7")
ACCENT_ROSE   = PatternFill("solid", fgColor="FEE2E2")
ACCENT_SKY    = PatternFill("solid", fgColor="E0F2FE")

wb = Workbook()
# remove default sheet, we'll add named ones
wb.remove(wb.active)


# =========================================================================
# Helpers
# =========================================================================
def style_header_row(ws, row_idx, num_cols):
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = CELL_BORDER
    ws.row_dimensions[row_idx].height = 32


def style_data_rows(ws, start_row, end_row, num_cols, banded=True):
    for r in range(start_row, end_row + 1):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = CELL_BORDER
            if cell.alignment is None or not cell.alignment.wrap_text:
                cell.alignment = LEFT_TOP
            if banded and (r - start_row) % 2 == 1:
                cell.fill = BAND_FILL
        ws.row_dimensions[r].height = 30


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_status_cf(ws, col_letter, first_row, last_row):
    """Apply status conditional formatting to a column range."""
    rng = f"{col_letter}{first_row}:{col_letter}{last_row}"
    rules = [
        ("Not Started", "F3F4F6", "374151"),
        ("In Progress", "FEF3C7", "92400E"),
        ("Complete",    "D1FAE5", "065F46"),
        ("Done",        "D1FAE5", "065F46"),
        ("Blocked",     "FEE2E2", "991B1B"),
        ("Skipped",     "E5E7EB", "4B5563"),
    ]
    for text, bg, fg in rules:
        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator="equal", formula=[f'"{text}"'],
                       fill=PatternFill("solid", fgColor=bg),
                       font=Font(color=fg, bold=True))
        )


def add_difficulty_cf(ws, col_letter, first_row, last_row):
    rng = f"{col_letter}{first_row}:{col_letter}{last_row}"
    rules = [
        ("Easy",   "D1FAE5", "065F46"),
        ("Medium", "FEF3C7", "92400E"),
        ("Hard",   "FEE2E2", "991B1B"),
    ]
    for text, bg, fg in rules:
        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator="equal", formula=[f'"{text}"'],
                       fill=PatternFill("solid", fgColor=bg),
                       font=Font(color=fg, bold=True))
        )


def add_databar(ws, col_letter, first_row, last_row):
    rng = f"{col_letter}{first_row}:{col_letter}{last_row}"
    rule = DataBarRule(start_type="num", start_value=0,
                       end_type="num", end_value=100,
                       color="6366F1", showValue=True)
    ws.conditional_formatting.add(rng, rule)


def add_validation(ws, col_letter, first_row, last_row, options):
    formula = '"' + ",".join(options) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.add(f"{col_letter}{first_row}:{col_letter}{last_row}")
    ws.add_data_validation(dv)


def write_title(ws, title, subtitle, last_col_letter):
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 36
    ws.merge_cells(f"A2:{last_col_letter}2")
    ws["A2"] = subtitle
    ws["A2"].font = SUBTITLE_FONT
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 20


# =========================================================================
# SHEET 2: Weekly Goals (build first so Dashboard can reference it)
# =========================================================================
ws_wg = wb.create_sheet("Weekly Goals")

weekly_data = [
    (1, "Foundations", "Setup & Python Mastery",
     "Set up a professional dev environment • Master Git/GitHub workflows • Refresh Python deeply with type hints • Get comfortable on Linux CLI • Learn pytest basics",
     "VS Code, Git, GitHub, Linux CLI, Python 3.11, type hints, virtualenv, pip, ruff, black, mypy, pytest",
     "Working CLI Task Manager + clean dev environment", 45),
    (2, "CS Fundamentals", "Core Computer Science",
     "Understand OS internals • Grasp networking from sockets to HTTP • Learn relational DB theory • Build intuition for indexes & transactions • Create 50 flashcards",
     "Processes, threads, memory, scheduling, TCP/IP, HTTP/2, DNS, sockets, SQL, normalization, indexes, transactions, ACID",
     "CS Fundamentals notes + 50 spaced-repetition flashcards", 50),
    (3, "Software Engineering", "Clean Code & Testing",
     "Internalize SOLID • Apply core design patterns • Write maintainable code • Master unit + integration testing • Design proper REST APIs",
     "SOLID, GoF patterns, clean code, pytest, mocking, TDD, REST design, API versioning, OpenAPI",
     "Refactored TODO API with 90%+ test coverage", 50),
    (4, "Backend Mastery", "FastAPI & Postgres",
     "Build production FastAPI services • Model with SQLAlchemy 2.0 • Validate with Pydantic v2 • Implement JWT auth • Add background tasks",
     "FastAPI, SQLAlchemy 2.0, Alembic, Pydantic v2, JWT, OAuth2, middleware, dependency injection, background tasks",
     "Production REST API with auth, migrations & docs", 55),
    (5, "DevOps & Cloud", "Docker, AWS, CI/CD",
     "Containerize multi-service apps • Deploy to AWS • Build CI/CD pipelines • Configure Nginx reverse proxy • Manage secrets properly",
     "Docker, docker-compose, multi-stage builds, AWS EC2, S3, RDS, Lambda, GitHub Actions, Nginx, env vars",
     "Dockerized app deployed live to AWS via CI/CD", 50),
    (6, "ML & LLM Foundations", "From Numpy to Transformers",
     "Refresh NumPy/Pandas • Train classical ML models • Build neural nets from scratch • Understand Transformer architecture • Use HuggingFace pipelines",
     "NumPy, Pandas, scikit-learn, PyTorch basics, MLPs, backprop, attention, Transformer blocks, HuggingFace, tokenization",
     "Sentiment Analysis API powered by HuggingFace", 55),
    (7, "RAG & Vector DBs", "Retrieval-Augmented Generation",
     "Master embeddings + chunking • Operate vector DBs • Implement hybrid search • Engineer prompts systematically • Evaluate RAG quality",
     "Embeddings, chunking strategies, Pinecone, Chroma, Qdrant, pgvector, hybrid search, prompt engineering, RAGAS",
     "Documentation Chatbot with production-grade RAG", 55),
    (8, "AI Agents", "LangGraph & Multi-Agent",
     "Build agents with LangChain • Orchestrate with LangGraph • Implement tool use • Understand MCP protocol • Compose multi-agent systems",
     "LangChain, LangGraph, ReAct, Plan-and-Execute, tool calling, MCP, agent memory, multi-agent orchestration",
     "Multi-Agent Research Assistant w/ web + code tools", 50),
    (9, "System Design & Production AI", "Scale & Serve",
     "Design scalable systems • Apply caching + queues • Serve LLMs in production • Add observability • Optimize cost & latency",
     "Scalability patterns, caching (Redis), queues (Kafka, SQS), microservices, vLLM, TGI, Prometheus, Grafana, cost ops",
     "LLM Gateway: rate-limited, cached, observable", 50),
    (10, "Capstone & Interviews", "Ship & Prepare",
     "Build end-to-end AI SaaS capstone • Polish portfolio site • Do 5 mock interviews • Refine resume • Prep behavioral stories",
     "Full-stack AI SaaS, Next.js, FastAPI, pgvector, agents, Stripe-mock, portfolio site, resume, STAR stories",
     "Capstone live + portfolio + 5 mock interviews done", 40),
]

wg_headers = ["Week #", "Phase Name", "Theme", "Primary Objectives", "Key Topics",
              "Deliverable / Milestone", "Target Hours", "Hours Logged",
              "Completion %", "Status", "Notes"]

write_title(ws_wg, "Weekly Goals", "Your 10-week phased curriculum — themes, objectives & deliverables", "K")
header_row = 4
for i, h in enumerate(wg_headers, 1):
    ws_wg.cell(row=header_row, column=i, value=h)
style_header_row(ws_wg, header_row, len(wg_headers))

for i, row in enumerate(weekly_data):
    r = header_row + 1 + i
    week_num = row[0]
    ws_wg.cell(row=r, column=1, value=row[0])
    ws_wg.cell(row=r, column=2, value=row[1])
    ws_wg.cell(row=r, column=3, value=row[2])
    ws_wg.cell(row=r, column=4, value=row[3])
    ws_wg.cell(row=r, column=5, value=row[4])
    ws_wg.cell(row=r, column=6, value=row[5])
    ws_wg.cell(row=r, column=7, value=row[6])
    # Hours Logged formula — SUMIF from Daily Tasks Week# column (B) and Actual Hours (J)
    ws_wg.cell(row=r, column=8,
               value=f"=SUMIF('Daily Tasks'!B:B,{week_num},'Daily Tasks'!J:J)")
    # Completion % formula — completed tasks / total tasks this week
    ws_wg.cell(row=r, column=9,
               value=(f'=IFERROR(COUNTIFS(\'Daily Tasks\'!B:B,{week_num},'
                      f'\'Daily Tasks\'!K:K,"Complete")/'
                      f'COUNTIF(\'Daily Tasks\'!B:B,{week_num}),0)'))
    ws_wg.cell(row=r, column=10, value="Not Started")
    ws_wg.cell(row=r, column=11, value="")

last_row = header_row + len(weekly_data)
style_data_rows(ws_wg, header_row + 1, last_row, len(wg_headers))

# Number formats
for r in range(header_row + 1, last_row + 1):
    ws_wg.cell(row=r, column=7).number_format = '0 "h"'
    ws_wg.cell(row=r, column=8).number_format = '0.0 "h"'
    ws_wg.cell(row=r, column=9).number_format = '0%'

set_col_widths(ws_wg, [8, 22, 26, 60, 55, 45, 12, 12, 12, 14, 30])
ws_wg.freeze_panes = "A5"
ws_wg.auto_filter.ref = f"A{header_row}:K{last_row}"

add_status_cf(ws_wg, "J", header_row + 1, last_row)
add_databar(ws_wg, "I", header_row + 1, last_row)
add_validation(ws_wg, "J", header_row + 1, last_row,
               ["Not Started", "In Progress", "Complete", "Blocked", "Skipped"])


# =========================================================================
# SHEET 3: Daily Tasks (70 rows, 7 days × 10 weeks)
# =========================================================================
ws_dt = wb.create_sheet("Daily Tasks")

# Real, granular tasks per day
daily_tasks = [
    # ---------- WEEK 1: Foundations ----------
    ("Foundations", "Dev Environment", "Install VS Code, Python 3.11 via pyenv; configure ruff, black, mypy; create dotfiles repo on GitHub", "Code", "Real Python: Setting up Python", 4.0),
    ("Foundations", "Git Basics", "Read Pro Git chapters 1-3; practice init/add/commit/branch/merge in scratch repo", "Reading", "https://git-scm.com/book", 5.0),
    ("Foundations", "Git Advanced", "Practice rebase, cherry-pick, interactive rebase, stash; resolve 3 merge conflicts intentionally", "Code", "Pro Git ch.7 + Atlassian Git tutorial", 5.0),
    ("Foundations", "Linux CLI", "Master bash basics: pipes, redirection, grep, sed, awk, find, xargs; build 5 one-liners", "Study", "MIT Missing Semester lec 1-3", 5.0),
    ("Foundations", "Python Refresh", "Fluent Python ch.1-3: data model, sequences, dict & set internals", "Reading", "Fluent Python — Luciano Ramalho", 6.0),
    ("Foundations", "Python Typing", "Master type hints, Protocol, TypedDict, Generic; convert an old script to fully typed + mypy strict clean", "Code", "mypy.readthedocs.io + PEP 484/544", 5.0),
    ("Foundations", "CLI Project", "Build CLI Task Manager with Click + SQLite + pytest; push to GitHub with README + CI workflow", "Project", "github.com/pallets/click", 6.0),

    # ---------- WEEK 2: CS Fundamentals ----------
    ("CS Fundamentals", "OS — Processes", "OSTEP ch.4-6: processes, API, scheduling; write notes + 10 flashcards", "Reading", "OSTEP (free PDF)", 6.0),
    ("CS Fundamentals", "OS — Threads", "OSTEP ch.26-28: threads, locks, condition variables; code a producer/consumer in Python", "Code", "OSTEP Concurrency", 6.0),
    ("CS Fundamentals", "OS — Memory", "Virtual memory, paging, TLB, page faults; draw memory layout diagram for a C/Python process", "Study", "OSTEP ch.13-22", 6.0),
    ("CS Fundamentals", "Networks — TCP/IP", "Beej's Networking Guide; explain 3-way handshake + TCP states from memory", "Reading", "beej.us/guide/bgnet", 6.0),
    ("CS Fundamentals", "Networks — HTTP", "HTTP/1.1 vs HTTP/2 vs HTTP/3; inspect real traffic with curl -v and Wireshark", "Code", "MDN HTTP docs", 6.0),
    ("CS Fundamentals", "DBMS — Relational", "Relational model, keys, normalization 1NF→3NF→BCNF; normalize a messy schema by hand", "Study", "CMU 15-445 lec 1-3", 5.0),
    ("CS Fundamentals", "DBMS — Indexes", "B-trees, hash indexes, query planning; run EXPLAIN ANALYZE on 5 queries in Postgres", "Code", "Use Mode SQL tutorial dataset", 5.0),

    # ---------- WEEK 3: Software Engineering ----------
    ("Software Engineering", "SOLID", "Read each SOLID principle with examples; refactor a smelly class to satisfy all 5", "Study", "Robert Martin — Clean Code", 5.0),
    ("Software Engineering", "Design Patterns", "Strategy, Factory, Observer, Repository, Adapter — implement each in Python", "Code", "refactoring.guru/design-patterns", 7.0),
    ("Software Engineering", "Clean Code", "Read Clean Code ch.2-5: naming, functions, comments, formatting; refactor an old project file", "Reading", "Clean Code — Robert Martin", 5.0),
    ("Software Engineering", "Testing — Unit", "pytest fixtures, parametrize, monkeypatch; achieve 90% coverage on CLI Task Manager", "Code", "pytest docs", 6.0),
    ("Software Engineering", "Testing — Integration", "Test DB layer with testcontainers; add Postman/httpx integration tests for the TODO API", "Code", "testcontainers.com", 6.0),
    ("Software Engineering", "REST Design", "Read REST API guidelines (Microsoft) + JSON:API; design resources for a blog API", "Study", "github.com/microsoft/api-guidelines", 5.0),
    ("Software Engineering", "Refactor Sprint", "Refactor TODO API: layered architecture, repository pattern, full test coverage; push v2 to GitHub", "Project", "Capstone of Week 3", 7.0),

    # ---------- WEEK 4: Backend Mastery ----------
    ("Backend Mastery", "FastAPI Core", "FastAPI tutorial 1-10: path/query/body params, response models, deps", "Reading", "fastapi.tiangolo.com", 6.0),
    ("Backend Mastery", "Postgres + SQLAlchemy", "Set up Postgres in Docker; model 4 tables with SQLAlchemy 2.0 declarative + relationships", "Code", "docs.sqlalchemy.org", 7.0),
    ("Backend Mastery", "Alembic Migrations", "Initialize Alembic; create 3 incremental migrations; practice downgrade", "Code", "alembic.sqlalchemy.org", 5.0),
    ("Backend Mastery", "Pydantic v2", "Master BaseModel, validators, computed fields, settings; migrate API schemas to v2", "Code", "docs.pydantic.dev", 6.0),
    ("Backend Mastery", "JWT Auth", "Implement /signup, /login, JWT access + refresh tokens with HTTP-only cookies", "Code", "FastAPI Security tutorial", 8.0),
    ("Backend Mastery", "Middleware & DI", "Custom middleware for logging + request ID; use Depends() for auth & DB session", "Code", "FastAPI Advanced", 6.0),
    ("Backend Mastery", "Project Build", "Wire everything: ship Production REST API with auth, migrations, OpenAPI docs, pytest suite", "Project", "Week 4 deliverable", 8.0),

    # ---------- WEEK 5: DevOps & Cloud ----------
    ("DevOps & Cloud", "Docker Basics", "Docker Deep Dive ch.1-5; write Dockerfile for FastAPI app with multi-stage build", "Reading", "Docker Deep Dive — Nigel Poulton", 6.0),
    ("DevOps & Cloud", "docker-compose", "Compose: api + postgres + redis + nginx; healthchecks; named volumes", "Code", "docs.docker.com/compose", 6.0),
    ("DevOps & Cloud", "AWS Basics", "AWS free tier setup; deploy EC2, S3 bucket, RDS Postgres; SSH + security groups", "Study", "AWS Free Tier docs", 6.0),
    ("DevOps & Cloud", "Deploy to EC2", "Deploy Dockerized API to EC2 behind Nginx reverse proxy with Let's Encrypt HTTPS", "Project", "Caddy/Nginx + certbot", 7.0),
    ("DevOps & Cloud", "CI/CD", "GitHub Actions: lint, test, build image, push to ECR, SSH-deploy to EC2", "Code", "docs.github.com/actions", 7.0),
    ("DevOps & Cloud", "Secrets & Env", "AWS Parameter Store / Secrets Manager; refactor app to pull secrets at runtime", "Code", "AWS Secrets Manager docs", 4.0),
    ("DevOps & Cloud", "Production Hardening", "Add logging (structlog), gunicorn workers, healthcheck endpoint, basic monitoring", "Code", "Week 5 deliverable", 6.0),

    # ---------- WEEK 6: ML & LLM Foundations ----------
    ("ML & LLM Foundations", "Numpy/Pandas Drills", "Solve 20 NumPy & 20 Pandas exercises (Stratascratch / Kaggle Learn)", "Practice", "kaggle.com/learn", 5.0),
    ("ML & LLM Foundations", "Classical ML", "Hands-On ML ch.3-4: classification, training; train Logistic Reg on Titanic + write up", "Reading", "Hands-On ML — Géron", 6.0),
    ("ML & LLM Foundations", "Neural Nets from Scratch", "Karpathy NN Zero to Hero lec 1-2: micrograd + MLP from scratch", "Video", "https://karpathy.ai", 8.0),
    ("ML & LLM Foundations", "PyTorch Basics", "PyTorch 60-min blitz; train MNIST MLP; understand autograd + DataLoader", "Code", "pytorch.org/tutorials", 6.0),
    ("ML & LLM Foundations", "Transformers", "Read 'Attention Is All You Need' + 'Illustrated Transformer'; sketch architecture from memory", "Reading", "arxiv.org/abs/1706.03762", 7.0),
    ("ML & LLM Foundations", "HuggingFace", "HF Course ch.1-3: pipeline, tokenizers, fine-tune DistilBERT on IMDb", "Code", "huggingface.co/learn", 8.0),
    ("ML & LLM Foundations", "Sentiment API", "Wrap fine-tuned model in FastAPI endpoint, dockerize, deploy to AWS", "Project", "Week 6 deliverable", 8.0),

    # ---------- WEEK 7: RAG & Vector DBs ----------
    ("RAG & Vector DBs", "Embeddings", "Read SBERT paper + OpenAI embeddings docs; compute & visualize embeddings with t-SNE", "Reading", "arxiv.org/abs/1908.10084", 5.0),
    ("RAG & Vector DBs", "Chunking", "Compare fixed, recursive, semantic chunking on a 200-page PDF; measure retrieval quality", "Code", "LangChain text splitters", 6.0),
    ("RAG & Vector DBs", "Vector DBs", "Spin up Chroma, Qdrant, pgvector; benchmark ingest + query on 100k docs", "Code", "docs.trychroma.com / qdrant.tech", 7.0),
    ("RAG & Vector DBs", "Hybrid Search", "Combine BM25 (rank_bm25) + dense retrieval; implement Reciprocal Rank Fusion", "Code", "Pinecone hybrid blog", 6.0),
    ("RAG & Vector DBs", "Prompt Engineering", "Read Lilian Weng prompt engineering post; build prompt library (system, few-shot, CoT, ReAct)", "Reading", "https://lilianweng.github.io", 6.0),
    ("RAG & Vector DBs", "RAG Evaluation", "Use RAGAS + custom LLM-as-judge; measure faithfulness, relevance, recall", "Code", "ragas.io", 6.0),
    ("RAG & Vector DBs", "Doc Chatbot", "Build Documentation Chatbot: ingest pipeline + Chroma + reranker + Streamlit UI; deploy", "Project", "Week 7 deliverable", 8.0),

    # ---------- WEEK 8: AI Agents ----------
    ("AI Agents", "LangChain Core", "LangChain docs: LCEL, runnables, chains, memory; rebuild RAG with LCEL", "Reading", "python.langchain.com", 6.0),
    ("AI Agents", "Tool Use", "Implement function-calling agent with 5 tools (web, math, files, sql, code); add tool errors handling", "Code", "OpenAI tool calling docs", 7.0),
    ("AI Agents", "LangGraph Basics", "LangGraph tutorial: nodes, edges, state; build a research agent with cycles", "Code", "langchain-ai.github.io/langgraph", 7.0),
    ("AI Agents", "ReAct vs Plan-Execute", "Implement both; benchmark on a 20-task suite; document tradeoffs", "Code", "arxiv.org/abs/2210.03629", 6.0),
    ("AI Agents", "Memory & State", "Long-term memory with vector store; conversation summary memory; thread persistence", "Code", "LangGraph checkpointers", 6.0),
    ("AI Agents", "MCP Protocol", "Read MCP spec; expose 3 tools via a local MCP server and connect from Claude Desktop", "Study", "modelcontextprotocol.io", 5.0),
    ("AI Agents", "Multi-Agent Project", "Build Multi-Agent Research Assistant (planner + researcher + writer) on LangGraph + Tavily", "Project", "Week 8 deliverable", 8.0),

    # ---------- WEEK 9: System Design & Production AI ----------
    ("System Design & Production AI", "SD Fundamentals", "Alex Xu Vol.1 ch.1-5: scaling, caching, sharding, message queues", "Reading", "System Design Interview — Alex Xu", 6.0),
    ("System Design & Production AI", "Caching Patterns", "Redis: cache-aside, write-through, TTL; implement semantic cache for LLM responses", "Code", "redis.io/docs", 6.0),
    ("System Design & Production AI", "Queues", "Kafka or SQS basics; build async LLM job queue with retries + DLQ", "Code", "kafka.apache.org / AWS SQS", 6.0),
    ("System Design & Production AI", "LLM Serving", "vLLM + TGI: continuous batching, paged attention; benchmark throughput vs naive HF", "Code", "github.com/vllm-project/vllm", 7.0),
    ("System Design & Production AI", "Observability", "Add Prometheus metrics + Grafana dashboards + structured logs + traces (OpenTelemetry)", "Code", "opentelemetry.io", 6.0),
    ("System Design & Production AI", "Cost & Latency", "Profile a RAG app: token cost, p50/p95 latency; cut cost 40% with caching + smaller models", "Practice", "Internal exercise", 5.0),
    ("System Design & Production AI", "LLM Gateway Project", "Build LLM Gateway: rate limit, semantic cache, model routing, metrics; deploy", "Project", "Week 9 deliverable", 8.0),

    # ---------- WEEK 10: Capstone & Interviews ----------
    ("Capstone & Interviews", "Capstone Scope", "Scope & design Capstone (AI SaaS): user stories, architecture diagram, DB schema, API plan", "Project", "Design doc in Notion", 6.0),
    ("Capstone & Interviews", "Capstone Build 1", "Build backend: FastAPI + Postgres + pgvector + auth + Stripe-mock + RAG pipeline", "Project", "Capstone repo", 8.0),
    ("Capstone & Interviews", "Capstone Build 2", "Build frontend: Next.js app with chat UI, dashboards, billing page; deploy to Vercel + AWS", "Project", "Capstone repo", 8.0),
    ("Capstone & Interviews", "Portfolio Site", "Build portfolio site (Next.js): projects, blog, resume; write 2 deep-dive blog posts", "Project", "Personal site live", 5.0),
    ("Capstone & Interviews", "Resume + LinkedIn", "Tailor resume + LinkedIn for AI SE roles; quantify every bullet; get 3 peer reviews", "Practice", "Levels.fyi resume guide", 4.0),
    ("Capstone & Interviews", "Mock Interviews", "Do 5 mock interviews (2 coding, 2 system design, 1 behavioral) on Pramp / interviewing.io", "Interview Prep", "interviewing.io", 5.0),
    ("Capstone & Interviews", "Behavioral Prep", "Prep 12 STAR stories covering leadership, conflict, failure, ambiguity, AI ethics; record yourself", "Interview Prep", "STAR method", 4.0),
]

assert len(daily_tasks) == 70, f"Expected 70 tasks, got {len(daily_tasks)}"

dt_headers = ["Day #", "Week #", "Date", "Phase", "Topic", "Task Description",
              "Task Type", "Resource / Reference", "Est. Hours", "Actual Hours",
              "Status", "Notes"]

write_title(ws_dt, "Daily Tasks", "Your 70-day battle plan — one row per day", "L")
header_row_dt = 4
for i, h in enumerate(dt_headers, 1):
    ws_dt.cell(row=header_row_dt, column=i, value=h)
style_header_row(ws_dt, header_row_dt, len(dt_headers))

for i, t in enumerate(daily_tasks):
    r = header_row_dt + 1 + i
    day_num = i + 1
    week_num = ((i) // 7) + 1
    ws_dt.cell(row=r, column=1, value=day_num)
    ws_dt.cell(row=r, column=2, value=week_num)
    ws_dt.cell(row=r, column=3, value=None)  # Date — user fills
    ws_dt.cell(row=r, column=4, value=t[0])
    ws_dt.cell(row=r, column=5, value=t[1])
    ws_dt.cell(row=r, column=6, value=t[2])
    ws_dt.cell(row=r, column=7, value=t[3])
    ws_dt.cell(row=r, column=8, value=t[4])
    ws_dt.cell(row=r, column=9, value=t[5])
    ws_dt.cell(row=r, column=10, value=0.0)
    ws_dt.cell(row=r, column=11, value="Not Started")
    ws_dt.cell(row=r, column=12, value="")

last_row_dt = header_row_dt + len(daily_tasks)
style_data_rows(ws_dt, header_row_dt + 1, last_row_dt, len(dt_headers))

for r in range(header_row_dt + 1, last_row_dt + 1):
    ws_dt.cell(row=r, column=3).number_format = 'YYYY-MM-DD'
    ws_dt.cell(row=r, column=9).number_format = '0.0 "h"'
    ws_dt.cell(row=r, column=10).number_format = '0.0 "h"'

set_col_widths(ws_dt, [7, 8, 12, 18, 22, 70, 14, 32, 10, 11, 13, 28])
ws_dt.freeze_panes = "A5"
ws_dt.auto_filter.ref = f"A{header_row_dt}:L{last_row_dt}"

add_status_cf(ws_dt, "K", header_row_dt + 1, last_row_dt)
add_validation(ws_dt, "K", header_row_dt + 1, last_row_dt,
               ["Not Started", "In Progress", "Complete", "Blocked", "Skipped"])
add_validation(ws_dt, "G", header_row_dt + 1, last_row_dt,
               ["Study", "Code", "Project", "Practice", "Revision", "Interview Prep", "Reading", "Video"])


# =========================================================================
# SHEET 1: Dashboard  (now that other sheets exist, we can reference)
# =========================================================================
ws_db = wb.create_sheet("Dashboard", 0)  # insert at position 0

write_title(ws_db, "AI Software Engineer OS — Dashboard",
            "Single source of truth for your 10-week journey to FAANG-level AI engineering",
            "H")

# KPI cards row 1 (rows 4-6 visual)
kpi_defs = [
    # (label, formula_or_value, fill, row, col_start)
    ("Overall Progress",
     "=IFERROR(COUNTIF('Daily Tasks'!K:K,\"Complete\")/70,0)", ACCENT_INDIGO),
    ("Current Week",
     "=MIN(10, INT((COUNTIF('Daily Tasks'!K:K,\"Complete\"))/7)+1)", ACCENT_SKY),
    ("Hours Logged",
     "=SUM('Daily Tasks'!J:J)", ACCENT_GREEN),
    ("Hours Target",
     500, ACCENT_AMBER),
    ("Tasks Complete",
     "=COUNTIF('Daily Tasks'!K:K,\"Complete\")", ACCENT_GREEN),
    ("Tasks Total",
     70, ACCENT_INDIGO),
    ("Streak (days)",
     0, ACCENT_ROSE),
    ("Interview Qs Solved",
     "=COUNTIFS('Interview Prep'!E:E,\"Complete\")", ACCENT_SKY),
]

# KPI cards with label + value structure
card_start_row = 4
card_width = 2   # 2 cells wide
card_height = 4  # 4 cells tall

for idx, (label, val, fill) in enumerate(kpi_defs):
    row_block = idx // 4
    col_block = idx % 4
    top = card_start_row + row_block * (card_height + 1)
    left = 1 + col_block * card_width

    # label row
    ws_db.merge_cells(start_row=top, start_column=left,
                      end_row=top, end_column=left + card_width - 1)
    lbl = ws_db.cell(row=top, column=left)
    lbl.value = label
    lbl.fill = fill
    lbl.font = Font(name="Calibri", size=10, bold=True, color="3730A3")
    lbl.alignment = Alignment(horizontal="center", vertical="center")
    lbl.border = Border(top=Side(border_style="medium", color="C7D2FE"),
                        left=Side(border_style="medium", color="C7D2FE"),
                        right=Side(border_style="medium", color="C7D2FE"))
    ws_db.row_dimensions[top].height = 22

    # value block
    ws_db.merge_cells(start_row=top + 1, start_column=left,
                      end_row=top + card_height - 1, end_column=left + card_width - 1)
    v = ws_db.cell(row=top + 1, column=left)
    v.value = val
    v.font = Font(name="Calibri", size=24, bold=True, color="1E1B4B")
    v.alignment = Alignment(horizontal="center", vertical="center")
    v.fill = WHITE_FILL
    v.border = Border(bottom=Side(border_style="medium", color="C7D2FE"),
                      left=Side(border_style="medium", color="C7D2FE"),
                      right=Side(border_style="medium", color="C7D2FE"))
    for rr in range(top + 1, top + card_height):
        ws_db.row_dimensions[rr].height = 22

    # number format
    if label == "Overall Progress":
        v.number_format = '0%'
    elif "Hours" in label:
        v.number_format = '0.0 "h"'

# Column widths for dashboard
set_col_widths(ws_db, [16, 16, 16, 16, 16, 16, 16, 16])

# Weekly Progress Summary table
wps_start = card_start_row + 2 * (card_height + 1) + 1
ws_db.merge_cells(start_row=wps_start, start_column=1, end_row=wps_start, end_column=8)
ws_db.cell(row=wps_start, column=1, value="Weekly Progress Summary").font = Font(
    name="Calibri", size=14, bold=True, color="1E1B4B")
ws_db.cell(row=wps_start, column=1).alignment = Alignment(horizontal="left", vertical="center")
ws_db.row_dimensions[wps_start].height = 26

wps_headers = ["Week", "Phase", "Status", "Completion %", "Hours Logged",
               "Hours Target", "Tasks Complete", "Tasks Total"]
hdr_row = wps_start + 1
for i, h in enumerate(wps_headers, 1):
    ws_db.cell(row=hdr_row, column=i, value=h)
style_header_row(ws_db, hdr_row, len(wps_headers))

for w in range(1, 11):
    r = hdr_row + w
    ws_db.cell(row=r, column=1, value=w)
    ws_db.cell(row=r, column=2, value=f"='Weekly Goals'!B{4 + w}")
    ws_db.cell(row=r, column=3, value=f"='Weekly Goals'!J{4 + w}")
    ws_db.cell(row=r, column=4, value=f"='Weekly Goals'!I{4 + w}")
    ws_db.cell(row=r, column=4).number_format = '0%'
    ws_db.cell(row=r, column=5, value=f"='Weekly Goals'!H{4 + w}")
    ws_db.cell(row=r, column=5).number_format = '0.0 "h"'
    ws_db.cell(row=r, column=6, value=f"='Weekly Goals'!G{4 + w}")
    ws_db.cell(row=r, column=6).number_format = '0 "h"'
    ws_db.cell(row=r, column=7, value=f'=COUNTIFS(\'Daily Tasks\'!B:B,{w},\'Daily Tasks\'!K:K,\"Complete\")')
    ws_db.cell(row=r, column=8, value=f'=COUNTIF(\'Daily Tasks\'!B:B,{w})')

style_data_rows(ws_db, hdr_row + 1, hdr_row + 10, len(wps_headers))
add_status_cf(ws_db, "C", hdr_row + 1, hdr_row + 10)
add_databar(ws_db, "D", hdr_row + 1, hdr_row + 10)

# Readiness Scores table
rs_start = hdr_row + 12
ws_db.merge_cells(start_row=rs_start, start_column=1, end_row=rs_start, end_column=4)
ws_db.cell(row=rs_start, column=1, value="Readiness Scores (0-100, self-assess weekly)").font = Font(
    name="Calibri", size=14, bold=True, color="1E1B4B")
ws_db.row_dimensions[rs_start].height = 26

rs_hdr = rs_start + 1
for i, h in enumerate(["Domain", "Score", "Level", "Notes"], 1):
    ws_db.cell(row=rs_hdr, column=i, value=h)
style_header_row(ws_db, rs_hdr, 4)

domains = ["Python", "SQL", "DSA", "System Design", "Backend / FastAPI",
           "Docker / Cloud", "ML Foundations", "LLM", "RAG", "AI Agents"]
for i, d in enumerate(domains):
    r = rs_hdr + 1 + i
    ws_db.cell(row=r, column=1, value=d)
    ws_db.cell(row=r, column=2, value=0)
    ws_db.cell(row=r, column=3, value=f'=IF(B{r}>=80,\"Strong\",IF(B{r}>=50,\"Growing\",IF(B{r}>=20,\"Learning\",\"Beginner\")))')
    ws_db.cell(row=r, column=4, value="")

style_data_rows(ws_db, rs_hdr + 1, rs_hdr + len(domains), 4)
add_databar(ws_db, "B", rs_hdr + 1, rs_hdr + len(domains))

ws_db.sheet_view.showGridLines = False
ws_db.freeze_panes = "A4"


# =========================================================================
# SHEET 4: Projects Tracker (10 projects, aligned to React app data)
# =========================================================================
ws_pj = wb.create_sheet("Projects Tracker")

projects = [
    (1, "Personal Finance Tracker (CLI)", 1, "Beginner", "Python · JSON", 1, 5, 0, "Not Started", 0,
     "github.com/yourname/cli-finance-tracker", "—",
     "OOP design classes (Transaction, Account), balance calculation, categories, local JSON read/write persistence, CSV export.",
     "Python classes inheritance, file system serialization, CLI menu event loop, CSV formatting.",
     "Built and shipped a command-line financial accountant tool in Python with JSON persistence and clean CSV export modules."),

    (2, "Browser History Simulator", 2, "Beginner", "Python", 1, 5, 0, "Not Started", 0,
     "github.com/yourname/history-simulator", "—",
     "LIFO browser navigation stack visit(), back(), forward() loops, history representation output logs.",
     "Stack operations push/pop, navigation pointers tracking, terminal log output formatting.",
     "Created an in-memory browser history navigation model utilizing multiple stacks in Python with undo/redo capabilities."),

    (3, "Expression AST Parser & Evaluator", 3, "Intermediate", "Python · AST", 3, 5, 0, "Not Started", 0,
     "github.com/yourname/ast-parser", "—",
     "Mathematical string parser utilizing Shunting-Yard infix to postfix conversion, AST nodes parsing, postorder traversal evaluator.",
     "Operator precedence parser mechanics, binary expression trees nodes creation, tree traversal recursive functions.",
     "Engineered a math expression parser from scratch in Python that parses strings into ASTs and evaluates them via postorder traversal."),

    (4, "Social Network Connection Router API", 4, "Intermediate", "Python · NetworkX · FastAPI", 3, 6, 0, "Not Started", 0,
     "github.com/yourname/social-router-api", "—",
     "Friends graph models using NetworkX adjacency lists, custom weighted Dijkstra paths search, mutual friends finder, FastAPI routers.",
     "Graph node/edge structures, Dijkstra shortest-path complexity optimization, REST API path params validation.",
     "Shipped a social network routing API using FastAPI and NetworkX that implements Dijkstra's algorithm to find friends pathways."),

    (5, "Task Management REST API (FastAPI + SQL)", 5, "Intermediate", "FastAPI · PostgreSQL · SQLAlchemy · Alembic", 3, 8, 0, "Not Started", 0,
     "github.com/yourname/fastapi-task-db-api", "—",
     "SQLAlchemy async models, Alembic migrations scripts, password hashing (bcrypt), JWT generation, authenticated task CRUD endpoints.",
     "Database async IO patterns, database migration loops, password encryption hashes, JWT tokens validation lifecycle.",
     "Built a secure task management REST API in FastAPI with PostgreSQL, SQLAlchemy ORM async engines, and JWT authorization checks."),

    (6, "Production Tasks Queue Service (Microservices)", 6, "Advanced", "FastAPI · Redis · Celery · Docker · Compose", 4, 10, 0, "Not Started", 0,
     "github.com/yourname/tasks-queue-service", "—",
     "Redis Cache-Aside backend query routing, custom sliding window rate limiters (Redis ZSET), Celery async workers task logs, Dockerfiles.",
     "Redis key TTL structures, sliding window logs transaction controls, Celery distributed tasks queuing, compose multi-container networks.",
     "Designed and containerized a scalable backend service incorporating Redis caching, Sliding-Window rate limiters, and Celery tasks queues."),

    (7, "Parking Lot Low-Level Design (OOD)", 7, "Intermediate", "Python · LLD · SOLID", 3, 6, 0, "Not Started", 0,
     "github.com/yourname/parking-lot-lld", "—",
     "SOLID compliant class structure mapping spot allocations, floor properties, payment gates, and ticket validation managers.",
     "Low-Level design principles, inheritance configurations, concurrency thread locks during concurrent checkout payments.",
     "Programmed a SOLID-compliant simulation of a multi-story parking lot with spot allocations logic and mock payment gates."),

    (8, "PyTorch Handwritten Digit Classifier", 8, "Advanced", "PyTorch · NumPy · Scikit-Learn", 4, 6, 0, "Not Started", 0,
     "github.com/yourname/mnist-pytorch-classifier", "—",
     "MNIST training loader pipeline, deep feedforward linear layers design, loss loops backpropagation steps, metrics precision report.",
     "Tensor manipulations, backprop mechanics, linear/nonlinear network configurations, classifier validation metrics calculation.",
     "Trained a handwritten digit classifier on PyTorch, implementing full dataset loading pipelines and metrics evaluations loops."),

    (9, "Semantic Document Q&A RAG Chatbot", 9, "Advanced", "FastAPI · ChromaDB · OpenAI API · Pytest", 4, 8, 0, "Not Started", 0,
     "github.com/yourname/rag-doc-chatbot", "—",
     "PDF text extraction, semantic paragraph splitting pipelines, ChromaDB vector store collection seeding, prompt injections.",
     "Vector databases index queries, semantic chunking strategy tradeoffs, OpenAI embeddings calls, integration testing (Pytest).",
     "Architected a Q&A chatbot using semantic RAG, integrating ChromaDB, OpenAI APIs, and validation modules verified via Pytest."),

    (10, "Flagship AI Research Assistant (Multi-Agent)", 10, "Capstone", "FastAPI · LangGraph · ChromaDB · React · Tavily API", 5, 20, 0, "Not Started", 0,
     "github.com/yourname/langgraph-research-agent", "https://research-assistant.demo.com",
     "LangGraph state machine supervisor nodes routing, Tavily web search tools integrations, server-sent events stream, React UI.",
     "Multi-agent loop states structures, streaming server-sent streams, React stream chunks parser, multi-container compose deploys.",
     "Launched a flagship multi-agent research portal utilizing LangGraph state control, stream parsing interfaces, and Tavily search tools.")
]

pj_headers = ["#", "Project Name", "Week", "Level", "Tech Stack", "Difficulty",
              "Est. Hours", "Hours Logged", "Status", "Progress %",
              "Repo Link", "Demo Link", "Key Features", "Skills Gained", "Resume Bullet"]

write_title(ws_pj, "Projects Tracker", "10 progressively harder projects — from CLI to production Multi-Agent AI SaaS", "O")
hr = 4
for i, h in enumerate(pj_headers, 1):
    ws_pj.cell(row=hr, column=i, value=h)
style_header_row(ws_pj, hr, len(pj_headers))

for i, p in enumerate(projects):
    r = hr + 1 + i
    for j, v in enumerate(p, 1):
        ws_pj.cell(row=r, column=j, value=v)

last_pj = hr + len(projects)
style_data_rows(ws_pj, hr + 1, last_pj, len(pj_headers))

for r in range(hr + 1, last_pj + 1):
    ws_pj.cell(row=r, column=7).number_format = '0 "h"'
    ws_pj.cell(row=r, column=8).number_format = '0.0 "h"'
    ws_pj.cell(row=r, column=10).number_format = '0%'
    ws_pj.cell(row=r, column=10).value = ws_pj.cell(row=r, column=10).value / 100 if isinstance(ws_pj.cell(row=r, column=10).value, (int, float)) else 0

set_col_widths(ws_pj, [5, 32, 7, 14, 42, 10, 11, 12, 13, 12, 38, 32, 55, 45, 70])
ws_pj.freeze_panes = "A5"
ws_pj.auto_filter.ref = f"A{hr}:O{last_pj}"

add_status_cf(ws_pj, "I", hr + 1, last_pj)
add_databar(ws_pj, "J", hr + 1, last_pj)
add_validation(ws_pj, "I", hr + 1, last_pj,
               ["Not Started", "In Progress", "Complete", "Blocked", "Skipped"])
add_validation(ws_pj, "D", hr + 1, last_pj,
               ["Beginner", "Intermediate", "Advanced", "Production", "Capstone"])


# =========================================================================
# SHEET 5: Milestones (20 milestones, 2 per week)
# =========================================================================
ws_ms = wb.create_sheet("Milestones")

milestones = [
    # week, name, desc, acceptance, reward
    (1, "Dev Environment Ready", "Fully configured local dev environment with linting and formatting on save", "VS Code, Python 3.11, ruff, black, mypy installed; dotfiles repo public", "New keyboard shortcut sticker on laptop ⌨️"),
    (1, "First Public CLI Shipped", "CLI Task Manager live on GitHub with CI passing", "Repo has README, MIT license, GH Actions green, >=80% test coverage", "Tweet the repo link 🐦"),
    (2, "CS Fundamentals Notes", "Complete OS + Networks + DBMS notes in knowledge base", ">=30 pages markdown, with diagrams and 50 flashcards", "Favorite coffee outing ☕"),
    (2, "First 50 Flashcards", "Seed your spaced-repetition system", "50 cards in Anki/Mochi/own tracker, all reviewed once", "1 movie night 🎬"),
    (3, "SOLID Refactor", "Refactor TODO API to layered architecture", "Repository pattern, services layer, 90%+ tests, no SRP violations", "Buy that programming book you wanted 📚"),
    (3, "First 100 LeetCode Easy/Medium", "Build problem-solving fluency", "100 problems solved on LeetCode, all neetcode.io easy patterns done", "Sunday off, no laptop 🌳"),
    (4, "Production REST API Live", "Ship the Week 4 deliverable", "API deployed locally with auth, migrations, OpenAPI; 25+ endpoints", "Treat yourself to a nice dinner 🍣"),
    (4, "Auth Mastery Checkpoint", "Explain JWT auth flow from memory on whiteboard", "5-min recorded explanation covering access/refresh/rotation/blacklist", "Share the recording on LinkedIn 💼"),
    (5, "First Container in Production", "Deploy first Docker container to AWS EC2", "Live URL serving HTTPS traffic through Nginx", "New mechanical keyboard? 😎"),
    (5, "CI/CD Pipeline Green", "Automated build/test/deploy pipeline", "Push to main → live in <5 min; rollback documented", "Weekend hike 🥾"),
    (6, "Neural Net from Scratch", "Build & train MLP without frameworks", "micrograd-style implementation; trains on MNIST to >=95%", "Frame the loss curve 📈"),
    (6, "Sentiment API Deployed", "HuggingFace model live in production", "Public endpoint with Swagger; >=85% accuracy on IMDb test", "Order a fun gadget 🎁"),
    (7, "First RAG Pipeline", "End-to-end RAG with eval", "Ingest → embed → retrieve → rerank → generate; RAGAS scores documented", "Game night 🎮"),
    (7, "Doc Chatbot Live", "Public Streamlit RAG app", "Anyone can chat with your docs; citations shown; eval report in README", "Post a demo video on Twitter/LinkedIn 📹"),
    (8, "ReAct Agent Working", "Tool-using agent on real tasks", "Solves >=80% of 20-task suite end-to-end with web + math tools", "New tech t-shirt 👕"),
    (8, "Multi-Agent System", "LangGraph multi-agent assistant", "Planner+Researcher+Writer produces 1500+ word research briefs", "Share architecture diagram on blog ✍️"),
    (9, "LLM Gateway in Production", "Serve real LLM traffic with limits + cache", "Rate limit + semantic cache + metrics live; load tested 100 RPS", "Premium coffee beans ☕"),
    (9, "System Design Mock Pass", "Pass a mock SD interview at 80%+", "Recorded mock + feedback from peer; covered scale/storage/caching", "Big weekend off 🏖️"),
    (10, "Capstone Shipped", "Capstone AI SaaS live and public", "End-to-end app live; signup → use feature → see billing; README full", "Celebration dinner 🥂"),
    (10, "FAANG-Ready", "Resume sent + 5 mock interviews done", "Resume reviewed; 5 mocks >=75% avg; 12 STAR stories memorized", "Apply to 10 target roles 🎯"),
]

ms_headers = ["#", "Week", "Milestone", "Description", "Acceptance Criteria",
              "Target Date", "Achieved Date", "Status", "Reward"]
write_title(ws_ms, "Milestones", "20 checkpoints that prove real progress — 2 per week", "I")
hr = 4
for i, h in enumerate(ms_headers, 1):
    ws_ms.cell(row=hr, column=i, value=h)
style_header_row(ws_ms, hr, len(ms_headers))

for i, m in enumerate(milestones):
    r = hr + 1 + i
    ws_ms.cell(row=r, column=1, value=i + 1)
    ws_ms.cell(row=r, column=2, value=m[0])
    ws_ms.cell(row=r, column=3, value=m[1])
    ws_ms.cell(row=r, column=4, value=m[2])
    ws_ms.cell(row=r, column=5, value=m[3])
    ws_ms.cell(row=r, column=6, value=None)
    ws_ms.cell(row=r, column=7, value=None)
    ws_ms.cell(row=r, column=8, value="Not Started")
    ws_ms.cell(row=r, column=9, value=m[4])

last_ms = hr + len(milestones)
style_data_rows(ws_ms, hr + 1, last_ms, len(ms_headers))
for r in range(hr + 1, last_ms + 1):
    ws_ms.cell(row=r, column=6).number_format = 'YYYY-MM-DD'
    ws_ms.cell(row=r, column=7).number_format = 'YYYY-MM-DD'

set_col_widths(ws_ms, [5, 7, 30, 45, 55, 14, 14, 14, 38])
ws_ms.freeze_panes = "A5"
ws_ms.auto_filter.ref = f"A{hr}:I{last_ms}"
add_status_cf(ws_ms, "H", hr + 1, last_ms)
add_validation(ws_ms, "H", hr + 1, last_ms,
               ["Not Started", "In Progress", "Complete", "Blocked", "Skipped"])


# =========================================================================
# SHEET 6: Interview Prep Tracker (60+ questions)
# =========================================================================
ws_iq = wb.create_sheet("Interview Prep")

iq = [
    # topic, question, difficulty, companies
    ("Python", "Explain the GIL and its impact on multithreading vs multiprocessing", "Medium", "Google, Meta"),
    ("Python", "What's the difference between __new__ and __init__? When would you override __new__?", "Medium", "Meta, Amazon"),
    ("Python", "Explain how Python's MRO works with multiple inheritance (C3 linearization)", "Hard", "Google, Stripe"),
    ("Python", "How does Python's garbage collector handle reference cycles?", "Hard", "Meta, Microsoft"),
    ("Python", "Implement a decorator that retries a function with exponential backoff", "Medium", "OpenAI, Anthropic"),
    ("Python", "What are metaclasses in Python and how do they modify class generation?", "Hard", "Google, Netflix"),
    ("Python", "Explain the difference between deepcopy and shallow copy in nested data structures", "Easy", "Apple, Microsoft"),
    ("Python", "How do generators work under the hood and what is the function of the yield statement?", "Medium", "Amazon, Uber"),

    ("SQL", "Write a query to find the second-highest salary per department", "Medium", "Amazon, Google"),
    ("SQL", "Explain B-tree vs hash indexes; when would you use each?", "Medium", "Meta, Stripe"),
    ("SQL", "What is a covering index? Show an EXPLAIN that benefits from one", "Hard", "Google, Snowflake"),
    ("SQL", "Difference between WHERE and HAVING? Can HAVING work without GROUP BY?", "Easy", "Amazon"),
    ("SQL", "Design a query to find users who logged in on N consecutive days", "Hard", "Meta, TikTok"),
    ("SQL", "Explain database normalization forms from 1NF to BCNF with design examples", "Medium", "Stripe, Square"),
    ("SQL", "Write a query utilizing window functions (e.g., LEAD, LAG, DENSE_RANK) to detect user sessions", "Hard", "Meta, Snowflake"),

    ("DSA", "Find longest substring without repeating characters", "Medium", "All FAANG"),
    ("DSA", "Implement LRU cache using a doubly linked list + hashmap", "Medium", "Amazon, Google"),
    ("DSA", "Word ladder (BFS) — minimum transformations between two words", "Hard", "Google, Meta"),
    ("DSA", "Top K frequent elements — discuss heap vs quickselect tradeoffs", "Medium", "Meta, OpenAI"),
    ("DSA", "Design and implement a thread-safe rate limiter (token bucket)", "Hard", "Stripe, Google"),
    ("DSA", "Reverse nodes in k-Group in a singly linked list in-place", "Hard", "Google, Apple"),
    ("DSA", "Find maximum path sum in a binary tree (any node to any node)", "Hard", "Facebook, ByteDance"),
    ("DSA", "Implement Dijkstra's shortest path algorithm using a min-heap", "Medium", "Microsoft, Amazon"),
    ("DSA", "Solve the 0/1 Knapsack problem with DP and print chosen items", "Medium", "Google, Bloomberg"),

    ("OOP", "Explain Liskov Substitution Principle with a code example", "Medium", "Microsoft, Amazon"),
    ("OOP", "Composition vs Inheritance — when do you favor each?", "Easy", "Meta"),
    ("OOP", "Design a parking lot system — classes, relationships, edge cases", "Hard", "Amazon, Google"),
    ("OOP", "Explain the Dependency Inversion Principle and how Dependency Injection enables it", "Medium", "Google, Spotify"),
    ("OOP", "Design a thread-safe Singleton class in Python using double-checked locking", "Medium", "Uber, Goldman Sachs"),

    ("OS", "Explain context switch and its cost", "Medium", "Google, Apple"),
    ("OS", "What is a deadlock? Show four conditions and a prevention strategy", "Medium", "Microsoft, Amazon"),
    ("OS", "Difference between mmap and read() for file I/O", "Hard", "Meta, Netflix"),
    ("OS", "Explain virtual memory paging and why page thrashing happens", "Medium", "Apple, NVIDIA"),
    ("OS", "What is the difference between a process and a thread in memory allocations?", "Easy", "All Tech"),

    ("DBMS", "What does ACID mean? Explain the trade-offs of each property in detail", "Medium", "Stripe, Oracle"),
    ("DBMS", "Explain MVCC (Multi-Version Concurrency Control) and how database engines avoid lock contentions", "Hard", "CockroachDB, Amazon"),
    ("DBMS", "Compare optimistic vs pessimistic locking and outline use cases for each", "Medium", "Stripe, Uber"),
    ("DBMS", "Explain horizontal partitioning (sharding) vs vertical partitioning and how to route queries", "Hard", "Databricks, Google"),

    ("System Design", "Design a URL shortener like Bitly (focus on database choice and base62 collision handling)", "Medium", "Twitter, Google"),
    ("System Design", "Design a real-time messaging app like Slack/WhatsApp (websockets vs polling, connection managers)", "Hard", "Meta, Telegram"),
    ("System Design", "Design an API rate limiter (token bucket, leaky bucket, redis sliding window configurations)", "Hard", "Stripe, Cloudflare"),
    ("System Design", "Explain consistent hashing rings and how they scale distributed storage clusters", "Hard", "Amazon, Netflix"),
    ("System Design", "Design a content delivery network (CDN) caching strategy and invalidation mechanisms", "Medium", "Fastly, Akamai"),
    ("System Design", "Compare WebSockets, Long Polling, and Server-Sent Events (SSE) with performance metrics", "Medium", "Twitch, Zoom"),

    ("AI/ML", "Explain the self-attention mechanism mathematically inside Transformer blocks", "Hard", "OpenAI, Anthropic"),
    ("AI/ML", "What is the structural difference between BERT (encoder-only) and GPT (decoder-only) architectures?", "Hard", "Google, Cohere"),
    ("AI/ML", "How does semantic chunking differ from simple recursive chunking in RAG pipelines?", "Medium", "Pinecone, LangChain"),
    ("AI/ML", "What are embeddings and why is cosine similarity preferred over Euclidean distance in high-dimensions?", "Easy", "OpenAI, Qdrant"),
    ("AI/ML", "Explain how RAGAS evaluations measure faithfulness and answer relevance in generation loops", "Medium", "HuggingFace, LlamaIndex"),
    ("AI/ML", "What are Multi-Agent topologies and how does LangGraph handle agent loops using state checkpointers?", "Hard", "LangChain, Microsoft"),

    ("Behavioral", "Tell me about a time you resolved a complex, critical production bug under intense time pressure", "Easy", "All Tech"),
    ("Behavioral", "Describe a technical disagreement you had with a senior developer and how you reached a consensus", "Easy", "All Tech"),
    ("Behavioral", "Tell me about a project that had highly ambiguous requirements and how you structured the scope", "Easy", "All Tech"),
    ("Behavioral", "Describe a major technical failure in a project you built and the recovery analysis steps taken", "Easy", "All Tech"),
    ("Behavioral", "Why are you looking to become an AI Software Engineer specifically, and what drives you in this field?", "Easy", "All Tech")
]

iq_headers = ["Topic", "Interview Question", "Difficulty", "Target Companies", "Status", "Notes"]
write_title(ws_iq, "Interview Prep Questions", "60+ core questions across Python, SQL, DSA, OOD, OS, DBMS, System Design, AI, and Behavioral", "F")

hr_iq = 4
for i, h in enumerate(iq_headers, 1):
    ws_iq.cell(row=hr_iq, column=i, value=h)
style_header_row(ws_iq, hr_iq, len(iq_headers))

for i, q in enumerate(iq):
    r = hr_iq + 1 + i
    ws_iq.cell(row=r, column=1, value=q[0])
    ws_iq.cell(row=r, column=2, value=q[1])
    ws_iq.cell(row=r, column=3, value=q[2])
    ws_iq.cell(row=r, column=4, value=q[3])
    ws_iq.cell(row=r, column=5, value="Not Started")
    ws_iq.cell(row=r, column=6, value="")

last_iq = hr_iq + len(iq)
style_data_rows(ws_iq, hr_iq + 1, last_iq, len(iq_headers))

set_col_widths(ws_iq, [18, 70, 14, 25, 16, 45])
ws_iq.freeze_panes = "A5"
ws_iq.auto_filter.ref = f"A{hr_iq}:F{last_iq}"

add_difficulty_cf(ws_iq, "C", hr_iq + 1, last_iq)
add_status_cf(ws_iq, "E", hr_iq + 1, last_iq)
add_validation(ws_iq, "E", hr_iq + 1, last_iq, ["Not Started", "In Progress", "Complete", "Blocked", "Skipped"])


# =========================================================================
# SHEET 7: LeetCode Tracker (100+ mapped problems)
# =========================================================================
ws_lc = wb.create_sheet("LeetCode Tracker")

lc_problems = [
    # Week 1
    (1, "Two Sum", "Easy", 1, "Arrays & Hash Map"),
    (9, "Palindrome Number", "Easy", 1, "Math"),
    (13, "Roman to Integer", "Easy", 1, "Strings & Hash Map"),
    (14, "Longest Common Prefix", "Easy", 1, "Strings"),
    (20, "Valid Parentheses", "Easy", 1, "Stack & Strings"),
    (66, "Plus One", "Easy", 1, "Arrays & Math"),
    (67, "Add Binary", "Easy", 1, "Strings & Math"),
    (83, "Remove Duplicates from Sorted List", "Easy", 1, "Linked List"),
    (88, "Merge Sorted Array", "Easy", 1, "Arrays & Two Pointers"),
    (118, "Pascal's Triangle", "Easy", 1, "Arrays & DP"),

    # Week 2
    (121, "Best Time to Buy and Sell Stock", "Easy", 2, "Arrays & Sliding Window"),
    (125, "Valid Palindrome", "Easy", 2, "Strings & Two Pointers"),
    (167, "Two Sum II", "Medium", 2, "Arrays & Two Pointers"),
    (169, "Majority Element", "Easy", 2, "Arrays"),
    (217, "Contains Duplicate", "Easy", 2, "Arrays & Hash Map"),
    (242, "Valid Anagram", "Easy", 2, "Strings & Hash Map"),
    (344, "Reverse String", "Easy", 2, "Strings & Two Pointers"),
    (349, "Intersection of Two Arrays", "Easy", 2, "Arrays & Hash Map"),
    (412, "Fizz Buzz", "Easy", 2, "Math"),
    (443, "String Compression", "Medium", 2, "Strings & Two Pointers"),
    (3, "Longest Substring Without Repeating Characters", "Medium", 2, "Strings & Sliding Window"),
    (11, "Container With Most Water", "Medium", 2, "Arrays & Two Pointers"),
    (15, "Three Sum", "Medium", 2, "Arrays & Two Pointers & Sorting"),
    (53, "Maximum Subarray", "Medium", 2, "Arrays & DP"),
    (128, "Longest Consecutive Sequence", "Medium", 2, "Arrays & Hash Map"),
    (49, "Group Anagrams", "Medium", 2, "Strings & Hash Map"),
    (202, "Happy Number", "Easy", 2, "Hash Map & Math"),
    (205, "Isomorphic Strings", "Easy", 2, "Strings & Hash Map"),
    (290, "Word Pattern", "Easy", 2, "Strings & Hash Map"),
    (347, "Top K Frequent Elements", "Medium", 2, "Arrays & Hash Map & Heap"),
    (383, "Ransom Note", "Easy", 2, "Hash Map & Strings"),
    (387, "First Unique Character in a String", "Easy", 2, "Hash Map & Strings"),
    (409, "Longest Palindrome", "Easy", 2, "Hash Map & Strings"),
    (70, "Climbing Stairs", "Easy", 2, "Recursion & DP"),
    (509, "Fibonacci Number", "Easy", 2, "Recursion & DP"),
    (1137, "N-th Tribonacci Number", "Easy", 2, "Recursion & DP"),

    # Week 3
    (206, "Reverse Linked List", "Easy", 3, "Linked List & Recursion"),
    (21, "Merge Two Sorted Lists", "Easy", 3, "Linked List"),
    (141, "Linked List Cycle", "Easy", 3, "Linked List & Two Pointers"),
    (160, "Intersection of Two Linked Lists", "Easy", 3, "Linked List & Two Pointers"),
    (203, "Remove Linked List Elements", "Easy", 3, "Linked List"),
    (234, "Palindrome Linked List", "Easy", 3, "Linked List & Two Pointers"),
    (237, "Delete Node in a Linked List", "Medium", 3, "Linked List"),
    (876, "Middle of the Linked List", "Easy", 3, "Linked List & Two Pointers"),
    (2, "Add Two Numbers", "Medium", 3, "Linked List & Math"),
    (19, "Remove Nth Node From End", "Medium", 3, "Linked List & Two Pointers"),
    (24, "Swap Nodes in Pairs", "Medium", 3, "Linked List & Recursion"),
    (92, "Reverse Linked List II", "Medium", 3, "Linked List"),
    (143, "Reorder List", "Medium", 3, "Linked List & Two Pointers"),
    (146, "LRU Cache", "Medium", 3, "Hash Map & Linked List"),
    (328, "Odd Even Linked List", "Medium", 3, "Linked List"),
    (23, "Merge K Sorted Lists", "Hard", 3, "Linked List & Heap"),
    (25, "Reverse Nodes in K-Group", "Hard", 3, "Linked List"),
    (155, "Min Stack", "Medium", 3, "Stack"),
    (739, "Daily Temperatures", "Medium", 3, "Stack & Monotonic Stack"),
    (84, "Largest Rectangle in Histogram", "Hard", 3, "Stack & Monotonic Stack"),
    (225, "Implement Stack Using Queues", "Easy", 3, "Stack & Queue"),
    (232, "Implement Queue Using Stacks", "Easy", 3, "Stack & Queue"),
    (622, "Design Circular Queue", "Medium", 3, "Queue"),
    (295, "Find Median from Data Stream", "Hard", 3, "Heap"),
    (703, "Kth Largest Element in a Stream", "Easy", 3, "Heap"),
    (496, "Next Greater Element I", "Easy", 3, "Stack & Monotonic Stack"),
    (503, "Next Greater Element II", "Medium", 3, "Stack & Monotonic Stack"),
    (901, "Online Stock Span", "Medium", 3, "Stack & Monotonic Stack"),
    (907, "Sum of Subarray Minimums", "Medium", 3, "Stack & Monotonic Stack & DP"),
    (150, "Evaluate Reverse Polish Notation", "Medium", 3, "Stack"),
    (394, "Decode String", "Medium", 3, "Stack & Strings"),
    (94, "Binary Tree Inorder Traversal", "Easy", 3, "Binary Tree & DFS"),
    (100, "Same Tree", "Easy", 3, "Binary Tree & DFS"),
    (101, "Symmetric Tree", "Easy", 3, "Binary Tree & DFS & BFS"),
    (104, "Maximum Depth of Binary Tree", "Easy", 3, "Binary Tree & DFS"),
    (111, "Minimum Depth of Binary Tree", "Easy", 3, "Binary Tree & BFS"),
    (112, "Path Sum", "Easy", 3, "Binary Tree & DFS"),
    (144, "Binary Tree Preorder Traversal", "Easy", 3, "Binary Tree & DFS"),
    (145, "Binary Tree Postorder Traversal", "Easy", 3, "Binary Tree & DFS"),
    (226, "Invert Binary Tree", "Easy", 3, "Binary Tree & DFS"),
    (235, "LCA of a BST", "Medium", 3, "BST & DFS"),
    (98, "Validate BST", "Medium", 3, "BST & DFS"),
    (105, "Construct BT from Preorder and Inorder", "Medium", 3, "Binary Tree & DFS"),
    (106, "Construct BT from Inorder and Postorder", "Medium", 3, "Binary Tree & DFS"),
    (114, "Flatten BT to Linked List", "Medium", 3, "Binary Tree & DFS"),
    (116, "Populating Next Right Pointers", "Medium", 3, "Binary Tree & BFS"),
    (173, "BST Iterator", "Medium", 3, "BST"),
    (199, "BT Right Side View", "Medium", 3, "Binary Tree & BFS"),
    (230, "Kth Smallest Element in a BST", "Medium", 3, "BST & DFS"),
    (236, "Lowest Common Ancestor of BT", "Medium", 3, "Binary Tree & DFS"),
    (108, "Convert Sorted Array to BST", "Easy", 3, "BST & Binary Search"),
    (124, "Binary Tree Maximum Path Sum", "Hard", 3, "Binary Tree & DFS"),
    (297, "Serialize and Deserialize BT", "Hard", 3, "Binary Tree & BFS & DFS"),

    # Week 4
    (200, "Number of Islands", "Medium", 4, "Graph & BFS/DFS"),
    (133, "Clone Graph", "Medium", 4, "Graph & BFS/DFS"),
    (1971, "Find if Path Exists in Graph", "Easy", 4, "Graph & BFS/DFS"),
    (695, "Max Area of Island", "Medium", 4, "Graph & DFS & Matrix"),
    (417, "Pacific Atlantic Water Flow", "Medium", 4, "Graph & DFS & Matrix"),
    (547, "Number of Provinces", "Medium", 4, "Graph & DFS & Union Find"),
    (207, "Course Schedule", "Medium", 4, "Graph & Topological Sort"),
    (210, "Course Schedule II", "Medium", 4, "Graph & Topological Sort"),
    (785, "Is Graph Bipartite", "Medium", 4, "Graph & BFS/DFS"),
    (743, "Network Delay Time", "Medium", 4, "Graph & Heap"),
    (787, "Cheapest Flights Within K Stops", "Medium", 4, "Graph & DP"),
    (1514, "Path with Maximum Probability", "Medium", 4, "Graph & Heap"),
    (684, "Redundant Connection", "Medium", 4, "Graph & Union Find"),
    (1319, "Number of Operations to Make Network Connected", "Medium", 4, "Graph & Union Find"),
    (310, "Minimum Height Trees", "Medium", 4, "Graph & BFS & Topological Sort"),
    (435, "Non-overlapping Intervals", "Medium", 4, "Greedy & Sorting"),
    (452, "Minimum Number of Arrows to Burst Balloons", "Medium", 4, "Greedy & Sorting"),
    (56, "Merge Intervals", "Medium", 4, "Sorting & Intervals"),
    (134, "Gas Station", "Medium", 4, "Greedy"),
    (621, "Task Scheduler", "Medium", 4, "Greedy & Heap"),

    # Week 5
    (198, "House Robber", "Medium", 5, "DP"),
    (213, "House Robber II", "Medium", 5, "DP"),
    (746, "Min Cost Climbing Stairs", "Easy", 5, "DP"),
    (740, "Delete and Earn", "Medium", 5, "DP"),
    (152, "Maximum Product Subarray", "Medium", 5, "DP & Arrays"),
    (45, "Jump Game II", "Medium", 5, "DP & Greedy"),
    (55, "Jump Game", "Medium", 5, "DP & Greedy"),
    (91, "Decode Ways", "Medium", 5, "DP & Strings"),
    (139, "Word Break", "Medium", 5, "DP & Strings"),
    (322, "Coin Change", "Medium", 5, "DP"),
    (377, "Combination Sum IV", "Medium", 5, "DP"),
    (62, "Unique Paths", "Medium", 5, "DP & Matrix"),
    (63, "Unique Paths II", "Medium", 5, "DP & Matrix"),
    (64, "Minimum Path Sum", "Medium", 5, "DP & Matrix"),
    (120, "Triangle", "Medium", 5, "DP"),
    (416, "Partition Equal Subset Sum", "Medium", 5, "DP"),
    (1049, "Last Stone Weight II", "Medium", 5, "DP"),
    (1143, "Longest Common Subsequence", "Medium", 5, "DP & Strings"),
    (72, "Edit Distance", "Medium", 5, "DP & Strings"),
    (583, "Delete Operation for Two Strings", "Medium", 5, "DP & Strings"),
    (5, "Longest Palindromic Substring", "Medium", 5, "DP & Strings"),
    (516, "Longest Palindromic Subsequence", "Medium", 5, "DP & Strings"),
    (647, "Palindromic Substrings", "Medium", 5, "DP & Strings"),
    (132, "Palindrome Partitioning II", "Hard", 5, "DP & Strings"),
    (312, "Burst Balloons", "Hard", 5, "DP & Intervals"),
    (10, "Regular Expression Matching", "Hard", 5, "DP & Strings"),

    # Week 6
    (78, "Subsets", "Medium", 6, "Backtracking & Arrays"),
    (90, "Subsets II", "Medium", 6, "Backtracking & Arrays"),
    (46, "Permutations", "Medium", 6, "Backtracking"),
    (47, "Permutations II", "Medium", 6, "Backtracking"),
    (77, "Combinations", "Medium", 6, "Backtracking"),
    (39, "Combination Sum", "Medium", 6, "Backtracking"),
    (40, "Combination Sum II", "Medium", 6, "Backtracking"),
    (216, "Combination Sum III", "Medium", 6, "Backtracking"),
    (22, "Generate Parentheses", "Medium", 6, "Backtracking & Strings"),
    (131, "Palindrome Partitioning", "Medium", 6, "Backtracking & Strings"),
    (93, "Restore IP Addresses", "Medium", 6, "Backtracking & Strings"),
    (51, "N-Queens", "Hard", 6, "Backtracking & Matrix"),
    (52, "N-Queens II", "Hard", 6, "Backtracking"),
    (37, "Sudoku Solver", "Hard", 6, "Backtracking & Matrix"),
    (79, "Word Search", "Medium", 6, "Backtracking & Matrix"),
    (212, "Word Search II", "Hard", 6, "Backtracking & Trie"),

    # Week 7
    (175, "Combine Two Tables", "Easy", 7, "SQL"),
    (176, "Second Highest Salary", "Medium", 7, "SQL"),
    (181, "Employees Earning More Than Their Managers", "Easy", 7, "SQL"),
    (182, "Duplicate Emails", "Easy", 7, "SQL"),
    (183, "Customers Who Never Order", "Easy", 7, "SQL"),
    (196, "Delete Duplicate Emails", "Easy", 7, "SQL"),
    (197, "Rising Temperature", "Easy", 7, "SQL"),
    (177, "Nth Highest Salary", "Medium", 7, "SQL"),
    (178, "Rank Scores", "Medium", 7, "SQL"),
    (184, "Department Highest Salary", "Medium", 7, "SQL"),
    (185, "Department Top Three Salaries", "Hard", 7, "SQL")
]

lc_headers = ["Week #", "Problem #", "Problem Title", "Difficulty", "Pattern / Category", "URL", "Status", "Date Solved", "Notes"]
write_title(ws_lc, "LeetCode Problems Tracker", "130+ LeetCode problems aligned week-by-week with your curriculum patterns", "I")

hr_lc = 4
for i, h in enumerate(lc_headers, 1):
    ws_lc.cell(row=hr_lc, column=i, value=h)
style_header_row(ws_lc, hr_lc, len(lc_headers))

for i, p in enumerate(lc_problems):
    r = hr_lc + 1 + i
    ws_lc.cell(row=r, column=1, value=p[3])
    ws_lc.cell(row=r, column=2, value=p[0])
    ws_lc.cell(row=r, column=3, value=p[1])
    ws_lc.cell(row=r, column=4, value=p[2])
    ws_lc.cell(row=r, column=5, value=p[4])
    ws_lc.cell(row=r, column=6, value=f"https://leetcode.com/problems/{p[1].lower().replace(' ', '-').replace(chr(39), '')}/")
    ws_lc.cell(row=r, column=7, value="Not Started")
    ws_lc.cell(row=r, column=8, value=None)
    ws_lc.cell(row=r, column=9, value="")

last_lc = hr_lc + len(lc_problems)
style_data_rows(ws_lc, hr_lc + 1, last_lc, len(lc_headers))

for r in range(hr_lc + 1, last_lc + 1):
    ws_lc.cell(row=r, column=8).number_format = 'YYYY-MM-DD'

set_col_widths(ws_lc, [8, 12, 35, 12, 28, 45, 15, 14, 30])
ws_lc.freeze_panes = "A5"
ws_lc.auto_filter.ref = f"A{hr_lc}:I{last_lc}"

add_difficulty_cf(ws_lc, "D", hr_lc + 1, last_lc)
add_status_cf(ws_lc, "G", hr_lc + 1, last_lc)
add_validation(ws_lc, "G", hr_lc + 1, last_lc, ["Not Started", "Attempted", "Complete", "Blocked"])


# =========================================================================
# SHEET 8: Resources Library
# =========================================================================
ws_res = wb.create_sheet("Resources Library")

resources = [
    # Week 1
    (1, "Python Official Tutorial", "Docs", "https://docs.python.org/3/tutorial", "Standard syntax guides, statements, and module systems."),
    (1, "Corey Schafer Python (Videos 1-20)", "YouTube", "https://youtube.com/c/CoreySchafer", "Essential Python configurations, scope rules, and library modules."),
    (1, "Fluent Python (Luciano Ramalho)", "Book", "https://www.oreilly.com/library/view/fluent-python-2nd/9781492056348/", "Pythonic protocols, data structures, and classes design models."),
    (1, "Pro Git (Scott Chacon & Ben Straub)", "Book", "https://git-scm.com/book", "Advanced branching model, local configurations, and internal mechanisms."),

    # Week 2
    (2, "OS: Three Easy Pieces (Arpaci-Dusseau)", "Book", "https://ostep.org", "Process cycles, threading synchronization locks, and CPU schedulers."),
    (2, "MIT Missing Semester (Lecture 1-3)", "Course", "https://missing.csail.mit.edu", "Bash scripting, pipeline structures, and tools configs."),
    (2, "Beej's Guide to Network Programming", "Docs", "https://beej.us/guide/bgnet", "Sockets bindings, connections handling, and TCP packets structures."),

    # Week 3
    (3, "Refactoring.Guru Design Patterns", "Docs", "https://refactoring.guru", "Design pattern definitions, Python code examples, and UML representations."),
    (3, "Clean Code (Robert C. Martin)", "Book", "https://www.oreilly.com/library/view/clean-code-a/9780136083238/", "SOLID properties, functions design, and names formatting rules."),

    # Week 4
    (4, "FastAPI Official Tutorial", "Docs", "https://fastapi.tiangolo.com", "Routing schemas, dependency injection models, and middleware controllers."),
    (4, "SQLAlchemy 2.0 Documentation", "Docs", "https://docs.sqlalchemy.org", "Declarative mappings, transaction contexts, and asynchronous databases engine calls."),

    # Week 5
    (5, "Docker Deep Dive (Nigel Poulton)", "Book", "https://dockerdeepdive.com", "Image layer optimizations, compose loops configurations, and docker networks."),
    (5, "AWS Free Tier Documentation", "Docs", "https://aws.amazon.com/free", "EC2, RDS Postgres provisioning, and S3 credentials configurations."),

    # Week 6
    (6, "Hands-On ML (Aurelien Geron)", "Book", "https://www.oreilly.com/library/view/hands-on-machine-learning/9781098125974/", "Machine learning metrics, regression bounds, and PyTorch properties."),
    (6, "Andrej Karpathy NN Zero to Hero", "YouTube", "https://karpathy.ai/zero-to-hero.html", "Neural net parameters calculations, backpropagation mathematical steps."),

    # Week 7
    (7, "LangChain Documentation", "Docs", "https://python.langchain.com", "LCEL model expressions, context documents loaders, and chains integrations."),
    (7, "ChromaDB Documentation", "Docs", "https://docs.trychroma.com", "Embeddings indexing properties and similarity search queries formats."),

    # Week 8
    (8, "LangGraph Documentation", "Docs", "https://langchain-ai.github.io/langgraph/", "Orchestrating agent state using cycles and persistent memory database savers."),
    (8, "Model Context Protocol Spec", "Docs", "https://modelcontextprotocol.io", "Exposing local server tools to client LLMs using MCP specifications."),

    # Week 9
    (9, "Designing Data-Intensive Applications", "Book", "https://dataintensive.net", "Consistency levels, partition sharding systems, and replica maps."),
    (9, "System Design Primer (Donne Martin)", "GitHub", "https://github.com/donnemartin/system-design-primer", "Scale models, consistent hashing configurations, CDNs caching maps."),

    # Week 10
    (10, "NeetCode System Design Course", "Course", "https://neetcode.io", "Architectural walk-throughs, scale evaluations, and interviewer mock logs.")
]

res_headers = ["Week #", "Resource Name", "Type", "Reference Link / URL", "Status", "Notes"]
write_title(ws_res, "Ranked Resources Library", "Top-tier reference books, official documentation, courses, and playlists", "F")

hr_res = 4
for i, h in enumerate(res_headers, 1):
    ws_res.cell(row=hr_res, column=i, value=h)
style_header_row(ws_res, hr_res, len(res_headers))

for i, r in enumerate(resources):
    row_idx = hr_res + 1 + i
    ws_res.cell(row=row_idx, column=1, value=r[0])
    ws_res.cell(row=row_idx, column=2, value=r[1])
    ws_res.cell(row=row_idx, column=3, value=r[2])
    ws_res.cell(row=row_idx, column=4, value=r[3])
    ws_res.cell(row=row_idx, column=5, value="Not Started")
    ws_res.cell(row=row_idx, column=6, value=r[4])

last_res = hr_res + len(resources)
style_data_rows(ws_res, hr_res + 1, last_res, len(res_headers))

set_col_widths(ws_res, [8, 35, 12, 45, 15, 60])
ws_res.freeze_panes = "A5"
ws_res.auto_filter.ref = f"A{hr_res}:F{last_res}"

add_status_cf(ws_res, "E", hr_res + 1, last_res)
add_validation(ws_res, "E", hr_res + 1, last_res, ["Not Started", "In Progress", "Complete"])
add_validation(ws_res, "C", hr_res + 1, last_res, ["Book", "YouTube", "Docs", "Course", "Interactive"])


# =========================================================================
# SHEET 9: Weekly Metrics Tracker
# =========================================================================
ws_met = wb.create_sheet("Weekly Metrics")

met_headers = [
    "Week #", "Study Hours (Target)", "Study Hours (Actual)",
    "LeetCode Problems (Target)", "LeetCode Problems (Actual)",
    "Mock Interview Completed", "Weak Areas Identified", "Corrective Actions"
]
write_title(ws_met, "Weekly Metrics Log", "Track targets vs. actual progress, revise weak topics weekly", "H")

hr_met = 4
for i, h in enumerate(met_headers, 1):
    ws_met.cell(row=hr_met, column=i, value=h)
style_header_row(ws_met, hr_met, len(met_headers))

for w in range(1, 11):
    r = hr_met + w
    ws_met.cell(row=r, column=1, value=w)
    ws_met.cell(row=r, column=2, value=50)  # 50 hours target
    ws_met.cell(row=r, column=3, value=f"='Weekly Goals'!H{4 + w}")  # Hours logged from Weekly Goals
    ws_met.cell(row=r, column=4, value=15)  # 15 problems target
    ws_met.cell(row=r, column=5, value=f"=COUNTIFS('LeetCode Tracker'!A:A,{w},'LeetCode Tracker'!G:G,\"Complete\")")
    ws_met.cell(row=r, column=6, value="No")
    ws_met.cell(row=r, column=7, value="")
    ws_met.cell(row=r, column=8, value="")

last_met = hr_met + 10
style_data_rows(ws_met, hr_met + 1, last_met, len(met_headers))

for r in range(hr_met + 1, last_met + 1):
    ws_met.cell(row=r, column=2).number_format = '0 "h"'
    ws_met.cell(row=r, column=3).number_format = '0.0 "h"'
    ws_met.cell(row=r, column=4).number_format = '0'
    ws_met.cell(row=r, column=5).number_format = '0'

set_col_widths(ws_met, [8, 20, 20, 22, 22, 22, 35, 35])
ws_met.freeze_panes = "A5"

add_validation(ws_met, "F", hr_met + 1, last_met, ["Yes", "No"])


# =========================================================================
# SHEET 10: Notes & Journal
# =========================================================================
ws_nt = wb.create_sheet("Notes & Journal")

nt_headers = ["Date", "Week #", "Day #", "Focus Area", "Key Takeaways / Concept Breakdown", "Code Snippets", "Blockers / Questions", "Next Steps"]
write_title(ws_nt, "Conceptual Notes & Daily Journal", "Log custom study notes, code snippets, blockages, and daily reflections", "H")

hr_nt = 4
for i, h in enumerate(nt_headers, 1):
    ws_nt.cell(row=hr_nt, column=i, value=h)
style_header_row(ws_nt, hr_nt, len(nt_headers))

for i in range(70):
    r = hr_nt + 1 + i
    day_num = i + 1
    week_num = (i // 7) + 1
    ws_nt.cell(row=r, column=1, value=None)
    ws_nt.cell(row=r, column=2, value=week_num)
    ws_nt.cell(row=r, column=3, value=day_num)
    ws_nt.cell(row=r, column=4, value=None)
    ws_nt.cell(row=r, column=5, value=None)
    ws_nt.cell(row=r, column=6, value=None)
    ws_nt.cell(row=r, column=7, value=None)
    ws_nt.cell(row=r, column=8, value=None)

last_nt = hr_nt + 70
style_data_rows(ws_nt, hr_nt + 1, last_nt, len(nt_headers), banded=True)

for r in range(hr_nt + 1, last_nt + 1):
    ws_nt.cell(row=r, column=1).number_format = 'YYYY-MM-DD'

set_col_widths(ws_nt, [12, 8, 8, 25, 45, 30, 25, 25])
ws_nt.freeze_panes = "A5"
ws_nt.auto_filter.ref = f"A{hr_nt}:H{last_nt}"


# =========================================================================
# SAVE WORKBOOK
# =========================================================================
file_path = os.path.join(r"d:\My Preparation", "AI_Software_Engineer_OS_Roadmap_10_Weeks.xlsx")
wb.save(file_path)
print(f"Workbook successfully saved to: {file_path}")
